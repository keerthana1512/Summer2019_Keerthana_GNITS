import openpyxl
import click
import os
from bs4 import BeautifulSoup
os.environ.setdefault('DJANGO_SETTINGS_MODULE','matchproject.settings')
import django
django.setup();
from matchapp.models import *

def get_data(sheet):
    data=[]
    for row in sheet.iter_rows():
        rowSelected = []
        for cell in row:
            rowSelected.append(cell.value)
        if not rowSelected[0]==None:
            data.append(rowSelected)
            print(rowSelected)
    return data


@click.group()
@click.option('-args')
def cli(args):
    pass

@cli.command(help="import data from excel to database using Django models")
def importdata():
    wb = openpyxl.load_workbook("matches.xlsx")
    ws = wb.get_sheet_by_name('matches')
    match_data=get_data(ws)[1:]
    for i in range(len(match_data)):
        print(match_data[i][0])
        m = Match(match_id=match_data[i][0],season=match_data[i][1],city=match_data[i][2],date=match_data[i][3],team1=match_data[i][4],team2=match_data[i][5],
                  toss_winner=match_data[i][6],toss_decision=match_data[i][7],result=match_data[i][8],dl_applied=match_data[i][9],winner=match_data[i][10],win_by_runs=match_data[i][11],
                  win_by_wickets=match_data[i][12],player_of_match=match_data[i][13],venue=match_data[i][14],umpire1=match_data[i][15],umpire2=match_data[i][16],umpire3=match_data[i][17])
        m.save()


@cli.command(help="import data from excel to database using Django models")
def importdata2():
    wb = openpyxl.load_workbook("deliveries.xlsx")
    ws = wb.get_sheet_by_name('deliveries')
    delivery_data=get_data(ws)[1:]
    for i in range(len(delivery_data)):
        print(delivery_data[i][0])
        d = Deliveries(m_id=Match.objects.get(match_id=delivery_data[i][0]),inning=delivery_data[i][1],
                       batting_team=delivery_data[i][2],
                       bowling_team=delivery_data[i][3],
                       over=delivery_data[i][4],
                       ball=delivery_data[i][5],
                       batsman=delivery_data[i][6],
                       non_striker=delivery_data[i][7],
                       bowler=delivery_data[i][8],
                       is_super_over=delivery_data[i][9],wide_runs=delivery_data[i][10],
                       bye_runs=delivery_data[i][11],legbye_runs=delivery_data[i][12],
                       noball_runs=delivery_data[i][13],penalty_runs=delivery_data[i][14],
                       batsman_runs=delivery_data[i][15],extra_runs=delivery_data[i][16],
                       total_runs=delivery_data[i][17],player_dismissed=delivery_data[i][18],
                       dismissal_kind=delivery_data[i][19],
                       fielder=delivery_data[i][20]
                       )
        d.save()


if __name__=='__main__':
    cli()
